import pandas as pd
import requests
import csv
import os
import sys,getopt
import time
import numpy as np
import re
import xlsxwriter
from openpyxl import load_workbook

def print_help():
    help_str=' python shopify2jd.py -v FNAME_JD -s FNAME_SHOPIFY [-o SKUOPTION] \n\
            -j FNAME_JD  (--JD=FNAME_JD) where FNAME_JD is the path to the JD template .csv file. Should just be this file: https://secure-wms.com/ViaSub.WMS/ImportFiles/Order%20Import%20Template.xlsx. The script will *TRY* to download this file but if the URL has changed, you will need to have a copy of it in your local area.\n\
            -s FNAME_SHOPIFY (--shopify==FNAME_SHOPIFY) where FNAME_SHOPIFY is the path to the SHOPIFY template .csv file\n\
            -o SKUOPTION (--only=SKUOPTION) can be "mats" "playpens" "balls" "accessories", filters out sku list from shopify and transforms them into a singular SKU.\n\n\
            For example, playpen-mat-balls-blue with "-only mats" option transforms it to "elite-play-mat-v2"\n\
            Another example, playpen-blue with "-only mats" option will just have the row dropped and ignored (won\'t be put into the JD output)'
    print(help_str)

def is_po_box(txt):
    po_box_detected = False
    if isinstance(txt,str):
        po_box_detected = re.search(r'(?:post(?:al)? (?:office )?|p[. ]?o\.? )?box',txt,re.IGNORECASE|re.MULTILINE)
    return True if po_box_detected else False

def process_args(opts, args):
    jd_fname = None
    shopify_fname = None
    only_sku_of = None
    for opt, arg in opts:
        if opt == "-h":
            print_help()
        elif opt in ("-j", "--jd"):
            jd_fname = arg
        elif opt in ("-s", "--shopify"):
            shopify_fname = arg
        elif opt in ("-o", "--only"):
            only_sku_of = arg
    return (jd_fname, shopify_fname, only_sku_of)

def build_items_per_order(items_per_order, curr_row_dict, name):
    if curr_row_dict['*Quantity'] > 1:
        print("WARNING FOR ORDER " + name + ": Quantity is more than 1 which means the shipping serivce generated will need to be changed! Surepost_under1lb for one mat only. Surepost_over1lb for over 10lbs only. UPS ground for >1lb, and <10lb")
    if name not in items_per_order:
        items_per_order[name] = [curr_row_dict['*Customer SKU ID']]
    else:
        items_per_order[name].append(curr_row_dict['*Customer SKU ID'])

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def shop2jd(shop_pd,jd_pd, only_sku_of, outfile):
#    for col in jd_pd['Outbound Order Info'].columns:
#        if "Outbound" in col:
#            print([col])
    customer_code = 'KH20000001629' #constant
    items_per_order = {}
    for idx, row in shop_pd.iterrows():
        if (row['Financial Status']!="paid"):
            # not paid yet / cancelled / refunded, so we're not gonna order anything for them.
            continue
        curr_row_dict = {}
        # convert all SKUs to the specified accessory
        # TODO_OPT - when i do a gui have a list of possible items to pick, like "elite-play-mat" "elite-play-mat-v2"

        #curr_row_dict is a dictionary that is a placeholder for the current shopify's .csv row being processed, expressed in terms of the destination .csv file
        curr_shop_sku = row['Lineitem sku']
        if only_sku_of=="mats":
            if "playpen-mat" in curr_shop_sku or "elite-play-mat" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = 'elite-play-mat-v2'
            else:
                continue #this order has no mat, so we don't need to fork the mat fulfillment to 3PL, skip this row.
        elif only_sku_of=="balls":
            if "playpen-mat-balls" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = 'pitballs-100' #only set of 100 balls comes with the playpen+mat+balls bundle
            elif "balls" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = curr_shop_sku #if they ordered balls by themselves.
            else:
                continue #not part of ppe+mat+balls bundle and customer didn't order balls by themselves, so there's no balls in this order to fork to 3PL.
        elif only_sku_of=="playpens":
            if "playpen-mat" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = "playpen-blue" if "blue" in curr_shop_sku else "playpen-red" #only 2 colors for now. #TODO_OPT - change this if >2 colors.
            elif "playpen" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = curr_shop_sku #not a bundle, so customer just ordered a playpen in which case we just use the SKU customer already has.
            else:
                continue #this is a non-playpen order and thus can be skipped.
        elif only_sku_of!="accessories":
            curr_row_dict['*Customer SKU ID'] = curr_shop_sku #no SKU filtering needed.


        curr_row_dict['*Customer Code'] = customer_code
        curr_row_dict['*Customer Order No.'] = row['Name']
        curr_row_dict['* Sales Channel No.'] =  "Playpen Elite" #constant, per sales channel
        curr_row_dict['*Sales Channel SO No.'] = row['Name'] # Same as customer order no.
        curr_row_dict['*Cargo Owner Code'] = 19916 #same as cargo owner ID 19916 or is this "Playpen Elite"?
        curr_row_dict['*Warehouse Code'] =  'C0000000579' #Fontana
        curr_row_dict['*Order Type\n1-B2C；2-B2B；\n3-WarehouseOnly'] = 1

        curr_row_dict['*Consignee Address1'] = row['Shipping Address1']
        curr_row_dict['Consignee Address2'] = row['Shipping Address2']
        #curr_row_dict['ShipToEmail'] = row['Email']
        curr_row_dict['*Consignee Name'] = row['Shipping Name'].lstrip().rstrip()
        #county needed?
        curr_row_dict['*Consignee City'] = row['Shipping City'].lstrip().rstrip()
        curr_row_dict['*Consignee State/Province'] = row['Shipping Province'].lstrip().rstrip()
        curr_row_dict['*Consignee Postcode'] = row['Shipping Zip']
        curr_row_dict['*Consignee Country'] = row['Shipping Country'].lstrip().rstrip()
        curr_row_dict['*Consignee District/County'] = "US"
        curr_row_dict['*Price'] = 1 # just a constant
        curr_row_dict['*Outbound Unit\n1-piece；2-Box；3-Pallet'] = 1 # Always 1 since we're b2c.

        phone = row['Shipping Phone']
        phone = phone.replace("(","").replace(")","").replace(" ","").replace("-","")
        if len(phone) == 11:
            phone = phone[1:]
        curr_row_dict['*Mobile'] = phone
        curr_row_dict['*Quantity'] = row['Lineitem quantity']
        if len(phone) <10 or len(phone)>11:
            print("MAKE SURE PHONE NUMBER FOR ORDER: " + row['Name'] + " is correct! It has " + str(len(phone)) + " digits!")

        # various warning checks for orders
        if (row['Risk Level']!="Low"):
            # TODO_OPT - just auto-drop medium or high risk orders??
            print("WARNING: Order " + row['Name'] + " has a \"" + row['Risk Level'] + "\" risk level! Please make sure this is something we want to import and not delete!")
        if (is_po_box(row['Shipping Address1']) or is_po_box(row['Shipping Address2'])):
            print("WARNING, ORDER " + row['Name'] + " HAS A PO BOX. WE CANNOT SHIP TO PO BOXES. PLEASE FIX THIS BEFORE IMPORTING IT TO JD!!!")

        curr_name = row['Name']
        if only_sku_of=="accessories": 
            if "playpen-mat" in curr_shop_sku or "elite-play-mat" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = 'elite-play-mat-v2' #there is at least a mat in the order
                jd_pd['Outbound Order Info'] = jd_pd['Outbound Order Info'].append(curr_row_dict, ignore_index=True)
                build_items_per_order(items_per_order, curr_row_dict, curr_name)
            if "playpen-mat-balls" in curr_shop_sku or "pitballs-100" in curr_shop_sku:
                curr_row_dict['*Customer SKU ID'] = 'pitballs-100' #there are balls in the order.
                jd_pd['Outbound Order Info'] = jd_pd['Outbound Order Info'].append(curr_row_dict, ignore_index=True)
                build_items_per_order(items_per_order, curr_row_dict, curr_name)
            continue # don't double push

        # build a dictionary of Order # -> Items to determine shipping service later.
        build_items_per_order(items_per_order, curr_row_dict, curr_name)

        jd_pd['Outbound Order Info'] = jd_pd['Outbound Order Info'].append(curr_row_dict, ignore_index=True)

    service_map = {
            "Surepost_under1lb" : "SPB9900hf",
            "Surepost_over1lb" : "SPB9900he",
            "UPS_ground" : "SPB99008u",
            }
    service_row = {}
    # work on second sheet, where we create shipping service based on weight
    for order in items_per_order:
        service_row['*Customer Code'] = customer_code
        service_row['*Customer Order No.'] = order
        items = items_per_order[order]
        if len(items) == 1:
            if 'elite-play-mat' in items[0]:
                # some mat, use Surepost_under1lb
                service_row['Service Product Code'] = service_map['Surepost_under1lb']
            elif 'playpen' in items[0]:
                # some playpen or playpen bundle, so >10lb and use Surepost_over1lb
                service_row['Service Product Code'] = service_map['Surepost_over1lb']
            else:
                # balls, so use UPS ground
                service_row['Service Product Code'] = service_map['UPS_ground']
        elif len(items) == 2:
            # either do UPS_ground or Surepost_over1lb (i.e. over 10lbs) as it is >1lb, and will be <10lb if no playpen is included or >=10lb if playpen is included
            if 'playpen-blue' in items or 'playpen-red' in items or 'playpen-mat-balls-blue' in items or 'playpen-mat-balls-red' in items or 'playpen-mat-blue' in items or 'playpen-mat-red' in items:
                # use Surepost_over1lb (>10lb)
                service_row['Service Product Code'] = service_map['Surepost_over1lb']
            elif ('elite-play-mat-v2' in items and 'pitballs-100' in items) or ('elite-play-mat' in items and 'pitballs-100' in items):
                # use UPS_ground (>1lb, but <10lb)
                service_row['Service Product Code'] = service_map['UPS_ground']
            else:
                print("Unexpected items for order: " + str(order) + " -- Items: " + str(items))
                sys.exit(2)
        else:
            #use UPS_Ground for simplicity
            service_row['Service Product Code'] = service_map['UPS_ground']
            print("Double check for order: " + str(order) + " to make sure that UPS ground is correct and it won't be over 10lbs (i.e. we should change it to Surepost_over1lb")

        jd_pd['Service Product Info'] = jd_pd['Service Product Info'].append(service_row, ignore_index=True)

    # need to open .xlsx file to write the ridiculous merge headers that JD put in.
    workbook = xlsxwriter.Workbook(outfile)
    outbound_order_info_ws = workbook.add_worksheet('Outbound Order Info')
    service_product_info_ws = workbook.add_worksheet('Service Product Info')
    merge_format = workbook.add_format({'align': 'center'})
    outbound_order_info_ws.merge_range('A1:L1', 'Outbound Order Info', merge_format)
    outbound_order_info_ws.merge_range('M1:Y1', 'Consignee Info', merge_format)
    outbound_order_info_ws.merge_range('Z1:AC1', 'Product Info', merge_format)
    service_product_info_ws.merge_range('A1:B1', 'Outbound Order Info', merge_format)
    service_product_info_ws.merge_range('C1:F1', 'Service Product Info', merge_format)

    workbook.close()


    #replace dataframe's NANs with empty string, as empty cells are NaNs
    jd_pd['Outbound Order Info']=jd_pd['Outbound Order Info'].replace(np.nan, '', regex=True)
    jd_pd['Service Product Info']=jd_pd['Service Product Info'].replace(np.nan, '', regex=True)

    append_df_to_excel(outfile, jd_pd['Outbound Order Info'], sheet_name='Outbound Order Info', startrow=1, index=False)
    append_df_to_excel(outfile, jd_pd['Service Product Info'], sheet_name='Service Product Info', startrow=1, index=False)
    #jd_pd['Outbound Order Info'].to_excel(writer,startrow=1,sheet_name='Outbound Order Info',index=False)
    #jd_pd['Service Product Info'].to_excel(writer,startrow=1,sheet_name='Service Product Info',index=False)
    
if __name__ == "__main__":
    # take in commandline arguments
    argv = sys.argv[1:]
    try:
        opts, args = getopt.getopt(argv,"hj:s:o:",["--jd=","--shopify=","--only="])
    except getopt.GetoptError:
        print_help()
        sys.exit(2)

    #run through each arg from commandline
    jd_fname, shopify_fname, only_sku_of= process_args(opts, args)

    # generate pd from jd and shopify data
    if (shopify_fname is None):
        print_help()
        print("--shopify= option is missing")
        sys.exit(2)
    if (jd_fname is None):
        print ("-j argument (--jd=) argument is missing. Please use -j to specify which .xlsx is the JD template file")
        sys.exit(2)
        # just use the -j argument for the template file.
    jd_pd = pd.read_excel(open(jd_fname,'rb'), sheet_name=None, header=1)
    #for some reason this imports a buncha unnamed columns for the first tab, so let's remove it
    cols_to_del = []
    for col in jd_pd['Outbound Order Info'].columns:
        if "Unnamed" in col:
            cols_to_del.append(col)
    if cols_to_del:
        jd_pd['Outbound Order Info'].drop(cols_to_del, axis=1, inplace=True) # don't bother entering this branch of cols_to_del is empty

    shop_pd = pd.read_csv(shopify_fname)
        
    # convert shopify to jd
    timestamp = time.strftime('%m-%d_%H%M%S', time.localtime())
    #TODO_OPT add a cmdline option for output file
    shop2jd(shop_pd=shop_pd,jd_pd=jd_pd, only_sku_of=only_sku_of, outfile="jd_"+timestamp+".xlsx")
    #TODO_OPT refactor this to be JD and VERDE orthogonally as options.
